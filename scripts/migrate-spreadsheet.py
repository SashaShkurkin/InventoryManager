#!/usr/bin/env python3
"""
Migrate data from 'Inventory.xlsx' into inventory_db.

Usage:
    pip install openpyxl
    python migrate-spreadsheet.py [--output migration.sql] [--db]

Flags:
    --output FILE   Write SQL INSERT statements to FILE (default: migration.sql)
    --db            Execute directly against MySQL (requires mysql-connector-python)
"""

import argparse
import re
import sys
from datetime import datetime, date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Install openpyxl first:  pip install openpyxl")

XLSX_PATH = Path(__file__).parent.parent / "Inventory.xlsx"

# Sheet name → state to assign
SHEET_STATE_MAP = {
    "Current Inventory": None,  # derived below
    "Inventory Archive": "Archived",
}
# Any sheet whose name contains "Sales" → Sold
SALES_SHEET_PATTERN = re.compile(r"sales", re.IGNORECASE)


def excel_date(value):
    """Convert Excel serial date number or datetime to ISO date string, or return None."""
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    try:
        n = float(value)
        # Excel epoch is 1899-12-30
        delta = datetime(1899, 12, 30) + __import__("datetime").timedelta(days=n)
        return delta.strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        return None


def clean_decimal(value):
    """Return a float or None from a cell value."""
    if value is None:
        return None
    s = str(value).replace("$", "").replace(",", "").strip()
    # Handle non-numeric values like "Bund." (bundled items)
    try:
        return float(s)
    except ValueError:
        return None


def escape(value):
    """SQL-escape a string value."""
    if value is None:
        return "NULL"
    return "'" + str(value).replace("'", "''").replace("\\", "\\\\") + "'"


def decimal_sql(value):
    v = clean_decimal(value)
    return str(v) if v is not None else "NULL"


def date_sql(value):
    d = excel_date(value)
    return f"'{d}'" if d else "NULL"


def parse_sheet(ws, default_state):
    """
    Return list of row dicts.  Column mapping is detected from the header row.
    """
    headers = []
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=5):
        for cell in row:
            if cell.value and str(cell.value).strip():
                # header row found
                header_row = row[0].row
                headers = [
                    str(c.value).strip().lower() if c.value else "" for c in row
                ]
                break
        if header_row:
            break

    if not header_row:
        return []

    # Flexible column matching
    def col(aliases):
        for alias in aliases:
            for i, h in enumerate(headers):
                if alias.lower() in h:
                    return i
        return None

    idx = {
        # SKU: fall back to column 0 if header is blank (Inventory Archive sheet)
        "sku":       col(["sku"]) if col(["sku"]) is not None else 0,
        "title":     col(["desc"]),
        "acq":       col(["acq"]),
        # "Added" in Archive/Sales sheets = labor cost
        "labor":     col(["labor", "added"]),
        "materials": col(["material"]),
        "prep":      col(["prep"]),
        "travel":    col(["travel"]),
        "shipping":  col(["shipping"]),
        "list":      col(["list"]),
        "sold":      col(["sold"]),
        "profit":    col(["profit"]),
        "type":      col(["type"]),
        "sub_type":  col(["sub"]),
        "style":     col(["style"]),
        "color":     col(["color"]),
        "date_acq":  col(["date acq"]),
        "date_sold": col(["date sold"]),
        "state_col": col(["state"]),
    }

    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        def v(key):
            i = idx.get(key)
            return row[i] if i is not None and i < len(row) else None

        sku = v("sku")
        if not sku or str(sku).strip() == "":
            continue

        title = v("title") or ""
        date_sold = v("date_sold")

        # Use explicit State column if present, otherwise derive
        state_val = v("state_col")
        if state_val and str(state_val).strip():
            state = str(state_val).strip().capitalize()
            # Normalize common variations
            state_map = {"sold": "Sold", "listed": "Listed", "processing": "Processing", "archived": "Archived"}
            state = state_map.get(state.lower(), state)
        elif default_state is None:
            state = "Sold" if excel_date(date_sold) else "Listed"
        else:
            state = default_state

        rows.append({
            "sku": str(sku).strip(),
            "title": str(title).strip(),
            "acq": v("acq"),
            "labor": v("labor"),
            "materials": v("materials"),
            "prep": v("prep"),
            "travel": v("travel"),
            "shipping": v("shipping"),
            "list_price": v("list"),
            "sold_price": v("sold"),
            "profit": v("profit"),
            "type": v("type"),
            "sub_type": v("sub_type"),
            "style": v("style"),
            "color": v("color"),
            "date_acq": v("date_acq"),
            "date_sold": date_sold,
            "state": state,
        })

    return rows


def generate_sql(all_rows):
    lines = [
        "-- Generated by migrate-spreadsheet.py",
        f"-- {datetime.now().isoformat()}",
        "",
        "SET NAMES utf8mb4;",
        "SET foreign_key_checks = 0;",
        "",
    ]

    seen_skus = set()
    for r in all_rows:
        sku = r["sku"]
        if sku in seen_skus:
            print(f"  [WARN] Duplicate SKU skipped: {sku}")
            continue
        seen_skus.add(sku)

        lines.append(
            "INSERT INTO InventoryItems "
            "(Sku, Title, AcquisitionCost, LaborCost, MaterialsCost, PrepCost, TravelCost, ShippingCost, "
            "ListPrice, SoldPrice, Profit, Type, SubType, Style, Color, "
            "DateAcquired, DateSold, State) VALUES ("
            f"{escape(sku)}, "
            f"{escape(r['title'])}, "
            f"{decimal_sql(r['acq'])}, "
            f"{decimal_sql(r['labor'])}, "
            f"{decimal_sql(r['materials'])}, "
            f"{decimal_sql(r['prep'])}, "
            f"{decimal_sql(r['travel'])}, "
            f"{decimal_sql(r['shipping'])}, "
            f"{decimal_sql(r['list_price'])}, "
            f"{decimal_sql(r['sold_price'])}, "
            f"{decimal_sql(r['profit'])}, "
            f"{escape(r['type'])}, "
            f"{escape(r['sub_type'])}, "
            f"{escape(r['style'])}, "
            f"{escape(r['color'])}, "
            f"{date_sql(r['date_acq'])}, "
            f"{date_sql(r['date_sold'])}, "
            f"{escape(r['state'])}"
            ");"
        )

    lines += ["", "SET foreign_key_checks = 1;", ""]
    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(description="Migrate Inventory.xlsx to SQL")
    parser.add_argument("--output", default="migration.sql")
    parser.add_argument("--db", action="store_true", help="Execute against MySQL directly")
    args = parser.parse_args()

    if not XLSX_PATH.exists():
        sys.exit(f"Spreadsheet not found: {XLSX_PATH}")

    print(f"Loading {XLSX_PATH} ...")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    all_rows = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if SALES_SHEET_PATTERN.search(sheet_name):
            state = "Sold"
        elif sheet_name in SHEET_STATE_MAP:
            state = SHEET_STATE_MAP[sheet_name]
        else:
            state = "Listed"

        rows = parse_sheet(ws, state)
        print(f"  Sheet '{sheet_name}': {len(rows)} rows  (state={state!r})")
        all_rows.extend(rows)

    print(f"\nTotal rows: {len(all_rows)}")
    sql = generate_sql(all_rows)

    out = Path(args.output)
    out.write_text(sql, encoding="utf-8")
    print(f"SQL written to: {out.resolve()}")

    if args.db:
        try:
            import mysql.connector  # type: ignore
        except ImportError:
            sys.exit("Install mysql-connector-python:  pip install mysql-connector-python")

        host = input("MySQL host [localhost]: ").strip() or "localhost"
        user = input("MySQL user [invuser]: ").strip() or "invuser"
        pw = input("MySQL password: ")
        db = input("Database [inventory_db]: ").strip() or "inventory_db"

        conn = mysql.connector.connect(host=host, user=user, password=pw, database=db)
        cursor = conn.cursor()
        for stmt in sql.split(";"):
            stmt = stmt.strip()
            if stmt and not stmt.startswith("--"):
                cursor.execute(stmt)
        conn.commit()
        cursor.close()
        conn.close()
        print("Done — data loaded into MySQL.")


if __name__ == "__main__":
    main()
